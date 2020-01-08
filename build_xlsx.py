from datetime import datetime

from openpyxl import Workbook, load_workbook


# def creating_xlsl():
#     wb = Workbook()
#     ws = wb.active
#
#     ws['B1'] = 'veikkaus'  # creating header for 'veikkaus data'
#     ws.merge_cells('B1:C1')
#     ws['B2'] = 'odds_1'
#     ws['C2'] = 'odds_2'
#
#     ws['D1'] = 'time'  # creating header for 'Time'
#     ws.merge_cells('D1:D2')
#
#     ws['E1'] = 'bet365'  # creating header for bet365 data'
#     ws.merge_cells('E1:H1')
#     ws['E2'] = 'odds_1'
#     ws['F2'] = 'odds_2'
#     ws['G2'] = 'col_1'
#     ws['H2'] = 'col_2'
#
#     ws['I1'] = 'William Hill'  # creating header for William Hill data
#     ws.merge_cells('I1:L1')
#     ws['I2'] = 'odds_1'
#     ws['J2'] = 'odds_2'
#     ws['K2'] = 'col_1'
#     ws['L2'] = 'col_2'
#
#     ws['M1'] = '1xBet'  # creating header for 1xBet data
#     ws.merge_cells('M1:P1')
#     ws['M2'] = 'odds_1'
#     ws['N2'] = 'odds_2'
#     ws['O2'] = 'col_1'
#     ws['P2'] = 'col_2'
#
#     ws['Q1'] = 'Pinnacle'  # creating header for Pinnacle data
#     ws.merge_cells('Q1:T1')
#     ws['Q2'] = 'odds_1'
#     ws['R2'] = 'odds_2'
#     ws['S2'] = 'col_1'
#     ws['T2'] = 'col_2'
#
#     timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
#     file_name = 'result-' + timestamp + '.xlsx'
#     wb.save(file_name)
#     return file_name


def check_and_handle_key(dictinary, key):
    if key in dictinary:
        return dictinary[key]
    else:
        return ' - '


def prpre_for_insrt_new_match(match):
    ready_list = list()
    for pos in range(22):
        ready_list.append(' - ')
    for item in match:
        for key in item:
            if key == 'veikkaus':
                ready_list[1] = check_and_handle_key(item[key], 'odds_1')
                ready_list[2] = check_and_handle_key(item[key], 'odds_2')
                # dummy for new matches
                ready_list[3] = check_and_handle_key(item[key], 'odds_1_new')
                ready_list[4] = check_and_handle_key(item[key], 'odds_2_new')
            elif key == 'time':
                ready_list[5] = check_and_handle_key(item, 'time')
            elif key == 'bet365':
                ready_list[6] = check_and_handle_key(item[key], 'odds_1')
                ready_list[7] = check_and_handle_key(item[key], 'odds_2')
                ready_list[8] = check_and_handle_key(item[key], 'col_1')
                ready_list[9] = check_and_handle_key(item[key], 'col_2')
            elif key == 'William Hill':
                ready_list[10] = check_and_handle_key(item[key], 'odds_1')
                ready_list[11] = check_and_handle_key(item[key], 'odds_2')
                ready_list[12] = check_and_handle_key(item[key], 'col_1')
                ready_list[13] = check_and_handle_key(item[key], 'col_2')
            elif key == '1xBet':
                ready_list[14] = check_and_handle_key(item[key], 'odds_1')
                ready_list[15] = check_and_handle_key(item[key], 'odds_2')
                ready_list[16] = check_and_handle_key(item[key], 'col_1')
                ready_list[17] = check_and_handle_key(item[key], 'col_2')
            elif key == 'Pinnacle':
                ready_list[18] = check_and_handle_key(item[key], 'odds_1')
                ready_list[19] = check_and_handle_key(item[key], 'odds_2')
                ready_list[20] = check_and_handle_key(item[key], 'col_1')
                ready_list[21] = check_and_handle_key(item[key], 'col_2')
    return ready_list


# def write_to_excel(matches_data):
#     file_name = creating_xlsl()
#     wb = load_workbook(file_name)
#     ws = wb.get_active_sheet()
#     checker = 3  # insert rows from 3 row. 2 first rows take header
#     for match in matches_data:
#         inserting_row = prpre_for_insrt_new_match(matches_data[match])
#         inserting_row[0] = match
#         ws.insert_rows(checker)
#         ws.append(inserting_row)
#         checker += 1
#     wb.save(file_name)


def prpr_for_insrt_exst_mtch(match):
    ready_list = list()
    for pos in range(22):
        ready_list.append(' - ')
    for item in match:
        for key in item:
            if key == 'veikkaus':
                # dummy for existing matches
                ready_list[1] = check_and_handle_key(item[key], 'odds_1_exist')
                ready_list[2] = check_and_handle_key(item[key], 'odds_2_exist')
                # end dummy for existing matches
                ready_list[3] = check_and_handle_key(item[key], 'odds_1')
                ready_list[4] = check_and_handle_key(item[key], 'odds_2')
            elif key == 'time':
                ready_list[5] = check_and_handle_key(item, 'time')
            elif key == 'bet365':
                ready_list[6] = check_and_handle_key(item[key], 'odds_1')
                ready_list[7] = check_and_handle_key(item[key], 'odds_2')
                ready_list[8] = check_and_handle_key(item[key], 'col_1')
                ready_list[9] = check_and_handle_key(item[key], 'col_2')
            elif key == 'William Hill':
                ready_list[10] = check_and_handle_key(item[key], 'odds_1')
                ready_list[11] = check_and_handle_key(item[key], 'odds_2')
                ready_list[12] = check_and_handle_key(item[key], 'col_1')
                ready_list[13] = check_and_handle_key(item[key], 'col_2')
            elif key == '1xBet':
                ready_list[14] = check_and_handle_key(item[key], 'odds_1')
                ready_list[15] = check_and_handle_key(item[key], 'odds_2')
                ready_list[16] = check_and_handle_key(item[key], 'col_1')
                ready_list[17] = check_and_handle_key(item[key], 'col_2')
            elif key == 'Pinnacle':
                ready_list[18] = check_and_handle_key(item[key], 'odds_1')
                ready_list[19] = check_and_handle_key(item[key], 'odds_2')
                ready_list[20] = check_and_handle_key(item[key], 'col_1')
                ready_list[21] = check_and_handle_key(item[key], 'col_2')
    return ready_list
