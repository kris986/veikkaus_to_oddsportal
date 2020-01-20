from datetime import datetime, timedelta

from openpyxl import Workbook, load_workbook

from build_xlsx import prpre_for_insrt_new_match
from build_xlsx import prpr_for_insrt_exst_mtch

from internal.oddsportal_base import OddsportalBase

from build_xlsx import check_and_handle_key

file_name = 'result.xlsx'
index_of_time = 6
oddsportal = OddsportalBase()


def creat_xlsl_file():
    wb = Workbook()
    ws = wb.active

    ws['B1'] = 'veikkaus'  # creating header for 'veikkaus data'
    ws.merge_cells('B1:F1')
    ws['B2'] = 'tournament'
    ws['C2'] = 'odds_1'
    ws['D2'] = 'odds_2'
    ws['E2'] = 'odds_1_new'
    ws['F2'] = 'odds_2_new'

    ws['G1'] = 'time'  # creating header for 'Time'
    ws.merge_cells('G1:G2')

    ws['H1'] = 'bet365'  # creating header for bet365 data'
    ws.merge_cells('H1:Q1')
    ws['H2'] = 'odds_1'
    ws['I2'] = 'odds_2'
    ws['J2'] = 'col_1'
    ws['K2'] = 'col_2'
    ws['L2'] = 'col_1_12h'
    ws['M2'] = 'col_2_12h'
    ws['N2'] = 'col_1_6h'
    ws['O2'] = 'col_2_6h'
    ws['P2'] = 'col_1_started'
    ws['Q2'] = 'col_2_started'

    ws['R1'] = 'William Hill'  # creating header for William Hill data
    ws.merge_cells('R1:AA1')
    ws['R2'] = 'odds_1'
    ws['S2'] = 'odds_2'
    ws['T2'] = 'col_1'
    ws['U2'] = 'col_2'
    ws['V2'] = 'col_1_12h'
    ws['W2'] = 'col_2_12h'
    ws['X2'] = 'col_1_6h'
    ws['Y2'] = 'col_2_6h'
    ws['Z2'] = 'col_1_started'
    ws['AA2'] = 'col_2_started'

    ws['AB1'] = '1xBet'  # creating header for 1xBet data
    ws.merge_cells('AB1:AK1')
    ws['AB2'] = 'odds_1'
    ws['AC2'] = 'odds_2'
    ws['AD2'] = 'col_1'
    ws['AE2'] = 'col_2'
    ws['AF2'] = 'col_1_12h'
    ws['AG2'] = 'col_2_12h'
    ws['AH2'] = 'col_1_6h'
    ws['AI2'] = 'col_2_6h'
    ws['AJ2'] = 'col_1_started'
    ws['AK2'] = 'col_2_started'

    ws['AL1'] = 'Pinnacle'  # creating header for Pinnacle data
    ws.merge_cells('AL1:AU1')
    ws['AL2'] = 'odds_1'
    ws['AM2'] = 'odds_2'
    ws['AN2'] = 'col_1'
    ws['AO2'] = 'col_2'
    ws['AP2'] = 'col_1_12h'
    ws['AQ2'] = 'col_2_12h'
    ws['AR2'] = 'col_1_6h'
    ws['AS2'] = 'col_2_6h'
    ws['AT2'] = 'col_1_started'
    ws['AU2'] = 'col_2_started'

    ws['AV1'] = 'result'  # creating header for 'Result of match'
    ws.merge_cells('AV1:AV2')

    wb.save(file_name)
    return file_name


def xlsl_file_exists():
    try:
        work_book = load_workbook(file_name)
        return work_book
    except FileNotFoundError:
        work_book = False
        return work_book


def update_xlsl_file(matches_data):
    work_book = xlsl_file_exists()
    if work_book:
        write_to_xlsl(work_book, matches_data)
    else:
        creat_xlsl_file()
        work_book = xlsl_file_exists()
        write_to_xlsl(work_book, matches_data)
    work_book.save(file_name)


def match_exist_in_sheet(ws, match):
    iter_rows = ws.iter_rows()
    for row in iter_rows:
        for cell in row:
            cell_val = cell.value
            if cell_val == match:
                num_row = cell.row
                return num_row
            else:
                num_row = False
    return num_row


def calc_delta_time(time):
    try:
        time = time.replace(',', '').replace(':', ' ')
        tm = datetime.strptime(time, '%d %b %Y %H %M')
        now = datetime.now()
        delta = (tm - now) / timedelta(minutes=1)
        delta = round((delta / 60), 2)
        return delta
    except ValueError:
        return 0


def update_row(ws, num_of_row, inserting_row):
    # for creating coordinates of cell
    cols_for_changing = list()
    cols = range(69, 92)
    for col in cols:
        if col == 91:
            a_cols = range(65, 87)
            for a_col in a_cols:
                a_cols_for_changing = [65, a_col]
                cols_for_changing.append(a_cols_for_changing)
        else:
            cols_for_changing.append(col)
    # updating Tournament_title
    ws[f'B{num_of_row}'].value = inserting_row[1]
    ind = 4
    for col in cols_for_changing:
        if ind < 7:
            ws[f'{chr(col)}{num_of_row}'].value = inserting_row[ind]
        elif ind in [9, 10, 19, 20, 29, 30, 39, 40]:
            # calculate time. return delta time
            delta = calc_delta_time(inserting_row[index_of_time])
            if 5.3 < delta < 6.8:
                col_index = cols_for_changing.index(col) + 4
                col = cols_for_changing[col_index]
                if isinstance(col, list):
                    ws[f'{chr(col[0])}{chr(col[1])}{num_of_row}'].value = inserting_row[ind]
                else:
                    ws[f'{chr(col)}{num_of_row}'].value = inserting_row[ind]
            elif 11.3 < delta < 12.8:
                col_index = cols_for_changing.index(col) + 2
                col = cols_for_changing[col_index]
                if isinstance(col, list):
                    ws[f'{chr(col[0])}{chr(col[1])}{num_of_row}'].value = inserting_row[ind]
                else:
                    ws[f'{chr(col)}{num_of_row}'].value = inserting_row[ind]

        if isinstance(col, list):
            if ws[f'{chr(col[0])}{chr(col[1])}{num_of_row}'].value == ' - ' and ind > 6:
                ws[f'{chr(col[0])}{chr(col[1])}{num_of_row}'].value = inserting_row[ind]
        else:
            if ws[f'{chr(col)}{num_of_row}'].value == ' - ' and ind > 6:
                ws[f'{chr(col)}{num_of_row}'].value = inserting_row[ind]
        ind += 1


def creating_new_matchs_row(ws, match_dict, match_name):
    inserting_row = prpre_for_insrt_new_match(match_dict)
    inserting_row[0] = match_name
    ws.append(inserting_row)


def write_to_xlsl(work_book, matches_data):
    timestamp = str(datetime.now().strftime("Last update: %d/%m/%Y %H:%M:%S"))
    ws = work_book.get_active_sheet()
    # inserting date of the last update
    ws['A3'].value = timestamp
    for match in matches_data:
        match_exists = match_exist_in_sheet(ws, match)  # (int) return number of row where match is written
        if match_exists:
            # calculate delta time. if past: create list of data as new match|prpre_for_insrt_new_match()
            delta = calc_delta_time(ws[f'G{match_exists}'].value)
            if delta < - 10:
                ws.delete_rows(match_exists)
                creating_new_matchs_row(ws, matches_data[match], match)
            else:
                inserting_row = prpr_for_insrt_exst_mtch(matches_data[match])
                inserting_row[0] = match
                update_row(ws, match_exists, inserting_row)
        else:
            creating_new_matchs_row(ws, matches_data[match], match)
    work_book.save(file_name)


def update_cell(ws, coordinate, value):
    ws[coordinate].value = value


def update_started_cells(ws, row, match):
    for item in match:
        for block in match[item]:
            for key in block:
                if key == 'bet365':
                    update_cell(ws, f'P{row}', check_and_handle_key(block[key], 'col_1'))
                    update_cell(ws, f'Q{row}', check_and_handle_key(block[key], 'col_2'))
                elif key == 'William Hill':
                    update_cell(ws, f'Z{row}', check_and_handle_key(block[key], 'col_1'))
                    update_cell(ws, f'AA{row}', check_and_handle_key(block[key], 'col_2'))
                elif key == '1xBet':
                    update_cell(ws, f'AJ{row}', check_and_handle_key(block[key], 'col_1'))
                    update_cell(ws, f'AK{row}', check_and_handle_key(block[key], 'col_2'))
                elif key == 'Pinnacle':
                    update_cell(ws, f'AT{row}', check_and_handle_key(block[key], 'col_1'))
                    update_cell(ws, f'AU{row}', check_and_handle_key(block[key], 'col_2'))


def analyze_existing_matches(driver):
    work_book = xlsl_file_exists()
    if work_book:
        ws = work_book.get_active_sheet()
        iter_rows = ws.iter_rows()
        ind_row = 0
        for row in iter_rows:
            if ind_row >= 3:
                for cell in row:
                    if cell.column == 7:
                        match_time = cell.value
                        match_dict = dict()
                        delta = calc_delta_time(match_time)
                        if -6 < delta < -0.01:
                            match_name = row[0].value
                            match_dict[match_name] = []
                            if -2 < delta < -0.01:
                                # for collecting coefficients after start match
                                match_dict = oddsportal.collect_data_by_dict(driver, match_dict)
                                # event_status = oddsportal.event_status(driver)
                                # if event_status == 'started':
                                update_started_cells(ws, cell.row, match_dict)

                            #  TODO prototype result posting to exl
                            elif -6 < delta <= -2:
                                oddsportal.try_searching(driver, match_name)
                                if oddsportal.handling_search_results_page(driver, match_name):
                                    result = oddsportal.collect_result_of_match(driver)
                                    if result:
                                        # print(f'{match_name}')
                                        update_cell(ws, f'AV{cell.row}', result)
            ind_row += 1
    work_book.save(file_name)
